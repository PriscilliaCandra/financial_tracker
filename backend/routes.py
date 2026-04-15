from flask import Blueprint, request, jsonify, g, send_file
from database import get_db
from auth import require_auth
from datetime import date, timedelta
import io

transactions_bp = Blueprint("transactions", __name__, url_prefix="/api")

def convert_to_idr(amount, currency):
    """Convert amount from given currency to IDR"""
    # Kurs sederhana (bisa diupdate atau panggil API)
    exchange_rates = {
        'IDR': 1,
        'USD': 15500,  # 1 USD = 15.500 IDR
        'SGD': 11500,  # 1 SGD = 11.500 IDR
        'MYR': 3300,   # 1 MYR = 3.300 IDR
        'EUR': 16800,  # 1 EUR = 16.800 IDR
        'JPY': 100,    # 1 JPY = 100 IDR
    }
    rate = exchange_rates.get(currency, 1)
    return round(amount * rate, 2)


# ── POST /api/transactions ────────────────────────────────────────────────────
@transactions_bp.route("/transactions", methods=["POST"])
@require_auth
def add_transaction():
    data = request.get_json() or {}
    required = ("date", "type", "category", "amount")
    missing = [f for f in required if not data.get(f)]
    if missing:
        return jsonify({"error": f"Missing fields: {', '.join(missing)}"}), 400
    if data["type"] not in ("income", "expense"):
        return jsonify({"error": "type must be 'income' or 'expense'"}), 400
    try:
        amount = float(data["amount"])
        if amount <= 0:
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({"error": "amount must be a positive number"}), 400

    # Ambil currency dari request, default IDR
    currency = data.get("currency", "IDR")
    amount = float(data["amount"])
    amount_idr = convert_to_idr(amount, currency)
    
    # Validasi currency (opsional)
    allowed_currencies = ["IDR", "USD", "SGD", "MYR", "EUR", "JPY"]
    if currency not in allowed_currencies:
        currency = "IDR"

    conn = get_db()
    cursor = conn.execute("""
        INSERT INTO transactions (user_id, date, type, category, amount, amount_idr, currency, note) 
        VALUES (?,?,?,?,?,?,?,?)
    """, (g.user_id, data["date"], data["type"], data["category"], amount, amount_idr, currency, data.get("note", "")))
    conn.commit()
    conn.close()
    return jsonify({"message": "Transaction added", "id": cursor.lastrowid}), 201


# ── GET /api/transactions ─────────────────────────────────────────────────────
@transactions_bp.route("/transactions", methods=["GET"])
@require_auth
def get_transactions():
    conn = get_db()
    query = "SELECT * FROM transactions WHERE user_id = ?"
    params = [g.user_id]

    if t := request.args.get("type"):
        query += " AND type = ?"
        params.append(t)
    if cat := request.args.get("category"):
        query += " AND category = ?"
        params.append(cat)
    if from_date := request.args.get("from"):
        query += " AND date >= ?"
        params.append(from_date)
    if to_date := request.args.get("to"):
        query += " AND date <= ?"
        params.append(to_date)

    query += " ORDER BY date DESC, created_at DESC"
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows]), 200


# ── PUT /api/transactions/<id> ────────────────────────────────────────────────
@transactions_bp.route("/transactions/<int:tx_id>", methods=["PUT"])
@require_auth
def edit_transaction(tx_id):
    conn = get_db()
    row = conn.execute(
        "SELECT id FROM transactions WHERE id = ? AND user_id = ?", (tx_id, g.user_id)
    ).fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "Transaction not found"}), 404

    data = request.get_json() or {}
    required = ("date", "type", "category", "amount")
    missing = [f for f in required if not data.get(f)]
    if missing:
        conn.close()
        return jsonify({"error": f"Missing fields: {', '.join(missing)}"}), 400
    if data["type"] not in ("income", "expense"):
        conn.close()
        return jsonify({"error": "type must be 'income' or 'expense'"}), 400
    try:
        amount = float(data["amount"])
        if amount <= 0:
            raise ValueError
    except (ValueError, TypeError):
        conn.close()
        return jsonify({"error": "amount must be a positive number"}), 400

    # Ambil currency dari request, default IDR
    currency = data.get("currency", "IDR")
    
    # Validasi currency (opsional)
    allowed_currencies = ["IDR", "USD", "SGD", "MYR", "EUR", "JPY"]
    if currency not in allowed_currencies:
        currency = "IDR"
        
    amount = float(data["amount"])
    amount_idr = convert_to_idr(amount, currency)

    conn.execute("""
        UPDATE transactions 
        SET date=?, type=?, category=?, amount=?, amount_idr=?, note=?, currency=? 
        WHERE id=? AND user_id=?
    """, (data["date"], data["type"], data["category"], amount, amount_idr, data.get("note", ""), currency, tx_id, g.user_id))
    
    
    conn.commit()
    conn.close()
    return jsonify({"message": "Transaction updated", "id": tx_id}), 200


# ── DELETE /api/transactions/<id> ─────────────────────────────────────────────
@transactions_bp.route("/transactions/<int:tx_id>", methods=["DELETE"])
@require_auth
def delete_transaction(tx_id):
    conn = get_db()
    row = conn.execute(
        "SELECT id FROM transactions WHERE id = ? AND user_id = ?", (tx_id, g.user_id)
    ).fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "Transaction not found"}), 404
    conn.execute("DELETE FROM transactions WHERE id = ? AND user_id = ?", (tx_id, g.user_id))
    conn.commit()
    conn.close()
    return jsonify({"message": "Transaction deleted", "id": tx_id}), 200


# ── GET /api/summary ──────────────────────────────────────────────────────────
@transactions_bp.route("/summary", methods=["GET"])
@require_auth
def get_summary():
    conn = get_db()
    query = """
        SELECT
            COALESCE(SUM(CASE WHEN type='income'  THEN amount_idr ELSE 0 END), 0) AS total_income,
            COALESCE(SUM(CASE WHEN type='expense' THEN amount_idr ELSE 0 END), 0) AS total_expense
        FROM transactions WHERE user_id = ?
    """
    params = [g.user_id]
    if from_date := request.args.get("from"):
        query += " AND date >= ?"
        params.append(from_date)
    if to_date := request.args.get("to"):
        query += " AND date <= ?"
        params.append(to_date)

    row = conn.execute(query, params).fetchone()
    conn.close()
    income = row["total_income"]
    expense = row["total_expense"]
    return jsonify({
        "total_income": income,
        "total_expense": expense,
        "balance": round(income - expense, 2),
    }), 200


# ── GET /api/salary-summary ───────────────────────────────────────────────────
@transactions_bp.route("/salary-summary", methods=["GET"])
@require_auth
def salary_summary():
    conn = get_db()
    today = date.today().isoformat()

    salary_rows = conn.execute("""
        SELECT id, date, amount, currency FROM transactions
        WHERE user_id = ? AND type = 'income' AND LOWER(category) = 'salary'
        ORDER BY date ASC, created_at ASC
    """, (g.user_id,)).fetchall()

    all_rows = conn.execute("""
        SELECT id, date, type, category, amount, note, currency FROM transactions
        WHERE user_id = ? ORDER BY date ASC, created_at ASC
    """, (g.user_id,)).fetchall()
    conn.close()

    all_txns = [dict(r) for r in all_rows]

    if not salary_rows:
        return jsonify({"cycles": [], "message": "No salary transactions found."}), 200

    salaries = [dict(r) for r in salary_rows]
    cycles = []
    for i, sal in enumerate(salaries):
        start = sal["date"]
        if i + 1 < len(salaries):
            end = (date.fromisoformat(salaries[i + 1]["date"]) - timedelta(days=1)).isoformat()
        else:
            end = today if today >= start else start
        cycles.append({
            "label": f"Cycle {i + 1}",
            "salary_date": start,
            "start": start,
            "end": end,
            "salary_amount": sal["amount"],
            "salary_currency": sal.get("currency", "IDR"),
            "is_current": (i + 1 == len(salaries)),
        })

    def find_cycle(tx_date):
        for idx, c in enumerate(cycles):
            if c["start"] <= tx_date <= c["end"]:
                return idx
        return -1

    buckets = {i: [] for i in range(len(cycles))}
    pre = []
    for tx in all_txns:
        idx = find_cycle(tx["date"])
        (buckets[idx] if idx >= 0 else pre).append(tx)

    def status(income, expense):
        if income <= 0:
            return "ok"
        ratio = (income - expense) / income
        return "critical" if ratio < 0.10 else "warning" if ratio < 0.30 else "ok"

    result = []
    for i, cyc in enumerate(cycles):
        txns = buckets[i]
        income = sum(t["amount_idr"] for t in txns if t["type"] == "income")
        expense = sum(t["amount_idr"] for t in txns if t["type"] == "expense")
        
        result.append({
            **cyc,
            "total_income": round(income, 2),
            "total_expense": round(expense, 2),
            "balance": round(income - expense, 2),
            "status": status(income, expense),
            "transactions": txns,
        })

    current = next((c for c in result if c["is_current"]), None)
    resp = {"cycles": result, "pre_salary": pre, "total_cycles": len(result)}
    if current:
        resp["current_cycle"] = {k: current[k] for k in
                                 ("label", "start", "end", "total_income", "total_expense", "balance", "status")}
    return jsonify(resp), 200


# ── GET /api/summary/daily ────────────────────────────────────────────────────
@transactions_bp.route("/summary/daily", methods=["GET"])
@require_auth
def daily_summary():
    """
    Returns income, expense, balance grouped by day.
    Optional: ?from=YYYY-MM-DD  ?to=YYYY-MM-DD
    """
    conn = get_db()
    query = """
        SELECT
            date,
            COALESCE(SUM(CASE WHEN type='income'  THEN amount_idr ELSE 0 END), 0) AS total_income,
            COALESCE(SUM(CASE WHEN type='expense' THEN amount_idr ELSE 0 END), 0) AS total_expense
        FROM transactions
        WHERE user_id = ?
    """
    params = [g.user_id]

    if from_date := request.args.get("from"):
        query += " AND date >= ?"
        params.append(from_date)
    if to_date := request.args.get("to"):
        query += " AND date <= ?"
        params.append(to_date)

    query += " GROUP BY date ORDER BY date DESC"
    rows = conn.execute(query, params).fetchall()
    conn.close()

    result = []
    for row in rows:
        income = row["total_income"]
        expense = row["total_expense"]
        result.append({
            "date": row["date"],
            "total_income": round(income, 2),
            "total_expense": round(expense, 2),
            "balance": round(income - expense, 2),
        })

    return jsonify(result), 200


# ── GET /api/summary/monthly ──────────────────────────────────────────────────
@transactions_bp.route("/summary/monthly", methods=["GET"])
@require_auth
def monthly_summary():
    """
    Returns income, expense, balance grouped by month (YYYY-MM).
    """
    conn = get_db()
    query = """
        SELECT
            strftime('%Y-%m', date) AS month,
            COALESCE(SUM(CASE WHEN type='income'  THEN amount_idr ELSE 0 END), 0) AS total_income,
            COALESCE(SUM(CASE WHEN type='expense' THEN amount_idr ELSE 0 END), 0) AS total_expense
        FROM transactions
        WHERE user_id = ?
    """
    params = [g.user_id]

    if from_date := request.args.get("from"):
        query += " AND date >= ?"
        params.append(from_date)
    if to_date := request.args.get("to"):
        query += " AND date <= ?"
        params.append(to_date)

    query += " GROUP BY month ORDER BY month DESC"
    rows = conn.execute(query, params).fetchall()
    conn.close()

    result = []
    for row in rows:
        income = row["total_income"]
        expense = row["total_expense"]
        result.append({
            "month": row["month"],
            "total_income": round(income, 2),
            "total_expense": round(expense, 2),
            "balance": round(income - expense, 2),
        })

    return jsonify(result), 200


# ── GET /api/budget?month=YYYY-MM ─────────────────────────────────────────────
@transactions_bp.route("/budget", methods=["GET"])
@require_auth
def get_budget():
    """
    Return the budget for a given month plus the current expense total.
    Defaults to the current month if ?month= is not provided.
    Also returns status: 'ok' | 'warning' (>=80%) | 'critical' (>=100%)
    """
    month = request.args.get("month") or date.today().strftime("%Y-%m")

    conn = get_db()

    # Fetch the limit (may not exist yet)
    budget_row = conn.execute(
        "SELECT limit_amount FROM budgets WHERE user_id = ? AND month = ?",
        (g.user_id, month)
    ).fetchone()

    # Fetch actual expense total for this month
    expense_row = conn.execute("""
        SELECT COALESCE(SUM(amount_idr), 0) AS total_expense
        FROM   transactions
        WHERE  user_id = ? AND type = 'expense'
          AND  strftime('%Y-%m', date) = ?
    """, (g.user_id, month)).fetchone()

    conn.close()

    limit = budget_row["limit_amount"] if budget_row else None
    expense = expense_row["total_expense"]

    # Calculate status only when a limit is set
    status = None
    pct = None
    if limit:
        pct = round((expense / limit) * 100, 1)
        status = "critical" if pct >= 100 else "warning" if pct >= 80 else "ok"

    return jsonify({
        "month": month,
        "limit_amount": limit,
        "total_expense": round(expense, 2),
        "remaining": round(limit - expense, 2) if limit else None,
        "percent_used": pct,
        "status": status,
    }), 200


# ── POST /api/budget ──────────────────────────────────────────────────────────
@transactions_bp.route("/budget", methods=["POST"])
@require_auth
def set_budget():
    """
    Create or update the expense limit for a month.
    Body: { "month": "YYYY-MM", "limit_amount": 3000000 }
    month defaults to current month if omitted.
    """
    data = request.get_json() or {}
    month = data.get("month") or date.today().strftime("%Y-%m")

    try:
        limit = float(data.get("limit_amount", 0))
        if limit <= 0:
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({"error": "limit_amount must be a positive number"}), 400

    conn = get_db()
    conn.execute("""
        INSERT INTO budgets (user_id, month, limit_amount)
        VALUES (?, ?, ?)
        ON CONFLICT(user_id, month) DO UPDATE SET limit_amount = excluded.limit_amount
    """, (g.user_id, month, limit))
    conn.commit()
    conn.close()

    return jsonify({"message": "Budget saved", "month": month, "limit_amount": limit}), 200


# ── DELETE /api/budget?month=YYYY-MM ─────────────────────────────────────────
@transactions_bp.route("/budget", methods=["DELETE"])
@require_auth
def delete_budget():
    """Remove the expense limit for a month."""
    month = request.args.get("month") or date.today().strftime("%Y-%m")

    conn = get_db()
    conn.execute(
        "DELETE FROM budgets WHERE user_id = ? AND month = ?",
        (g.user_id, month)
    )
    conn.commit()
    conn.close()
    return jsonify({"message": "Budget removed", "month": month}), 200


# ── GET /api/analytics ────────────────────────────────────────────────────────
@transactions_bp.route("/analytics", methods=["GET"])
@require_auth
def analytics():
    """
    Returns two datasets for charting:

    1. category_breakdown — expense totals grouped by category for the
       requested month (default = current month).

    2. monthly_trend — income and expense totals for each of the last
       N months (default 6), ordered oldest → newest for chart rendering.

    Query params:
      ?month=YYYY-MM   month for category breakdown  (default: current)
      ?months=N        how many months back for trend (default: 6, max: 24)
    """
    conn = get_db()
    today = date.today()

    # ── Param: which month for breakdown ─────────────────────
    month = request.args.get("month") or today.strftime("%Y-%m")

    # ── Param: how many months for trend ─────────────────────
    try:
        n_months = min(int(request.args.get("months", 6)), 24)
    except (ValueError, TypeError):
        n_months = 6

    # ── 1. Category breakdown (expenses only) ─────────────────
    cat_rows = conn.execute("""
        SELECT category,
               SUM(amount) AS total
        FROM   transactions
        WHERE  user_id = ?
          AND  type    = 'expense'
          AND  strftime('%Y-%m', date) = ?
        GROUP  BY category
        ORDER  BY total DESC
    """, (g.user_id, month)).fetchall()

    total_expense = sum(r["total"] for r in cat_rows)
    breakdown = []
    for row in cat_rows:
        breakdown.append({
            "category": row["category"],
            "amount": round(row["total"], 2),
            "percent": round((row["total"] / total_expense * 100), 1) if total_expense else 0,
        })

    # ── 2. Monthly trend (last N months) ──────────────────────
    # Build the list of YYYY-MM strings we want, oldest first
    trend_months = []
    for i in range(n_months - 1, -1, -1):
        y = today.year
        m = today.month - i
        while m <= 0:
            m += 12
            y -= 1
        trend_months.append(f"{y:04d}-{m:02d}")

    trend_rows = conn.execute("""
        SELECT strftime('%Y-%m', date)                                    AS month,
               COALESCE(SUM(CASE WHEN type='income'  THEN amount_idr END), 0) AS income,
               COALESCE(SUM(CASE WHEN type='expense' THEN amount_idr END), 0) AS expense
        FROM   transactions
        WHERE  user_id = ?
          AND  strftime('%Y-%m', date) >= ?
        GROUP  BY month
        ORDER  BY month ASC
    """, (g.user_id, trend_months[0])).fetchall()

    # Index by month so missing months get zeros
    trend_index = {r["month"]: r for r in trend_rows}
    trend = []
    for m in trend_months:
        row = trend_index.get(m)
        trend.append({
            "month": m,
            "income": round(row["income"], 2) if row else 0,
            "expense": round(row["expense"], 2) if row else 0,
        })

    conn.close()
    return jsonify({
        "month": month,
        "category_breakdown": breakdown,
        "total_expense": round(total_expense, 2),
        "monthly_trend": trend,
    }), 200


# ── GET /api/summary/categories ──────────────────────────────────────────────
@transactions_bp.route("/summary/categories", methods=["GET"])
@require_auth
def category_breakdown():
    """
    Returns expense totals grouped by category, with percentage share.
    Optional: ?month=YYYY-MM  (defaults to all time)
              ?from=YYYY-MM-DD  ?to=YYYY-MM-DD
    """
    conn = get_db()
    query = """
        SELECT
            category,
            COALESCE(SUM(amount_idr), 0) AS total
        FROM transactions
        WHERE user_id = ? AND type = 'expense'
    """
    params = [g.user_id]

    if month := request.args.get("month"):
        query += " AND strftime('%Y-%m', date) = ?"
        params.append(month)
    else:
        if from_date := request.args.get("from"):
            query += " AND date >= ?"
            params.append(from_date)
        if to_date := request.args.get("to"):
            query += " AND date <= ?"
            params.append(to_date)

    query += " GROUP BY category ORDER BY total DESC"
    rows = conn.execute(query, params).fetchall()
    conn.close()

    grand_total = sum(r["total"] for r in rows)

    result = []
    for row in rows:
        result.append({
            "category": row["category"],
            "total": round(row["total"], 2),
            "percent": round((row["total"] / grand_total * 100), 1) if grand_total else 0,
        })

    return jsonify({
        "categories": result,
        "grand_total": round(grand_total, 2),
    }), 200


# ── EXCEL & PDF EXPORT ─────────────────────────────────────────────────────
@transactions_bp.route("/export/excel", methods=["GET"])
@require_auth
def export_excel():
    """Export all user transactions to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    conn = get_db()
    rows = conn.execute("""
        SELECT date, type, category, amount, note, currency, created_at
        FROM transactions
        WHERE user_id = ?
        ORDER BY date DESC, created_at DESC
    """, (g.user_id,)).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_alignment = Alignment(horizontal="center")

    headers = ["Date", "Type", "Category", "Amount", "Currency", "Note", "Created At"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows
    for row in rows:
        ws.append([
            row["date"],
            row["type"].capitalize(),
            row["category"],
            row["amount"],
            row["currency"],
            row["note"] or "",
            row["created_at"]
        ])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name=f"transactions_{g.username}.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@transactions_bp.route("/export/pdf", methods=["GET"])
@require_auth
def export_pdf():
    """Export all user transactions to PDF file"""
    try:
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        return jsonify({"error": "reportlab not installed. Run: pip install reportlab"}), 500

    from datetime import datetime

    conn = get_db()
    rows = conn.execute("""
        SELECT date, type, category, amount, note, currency
        FROM transactions
        WHERE user_id = ?
        ORDER BY date DESC, created_at DESC
    """, (g.user_id,)).fetchall()
    conn.close()

    # Get summary totals
    total_income = sum(r["amount"] for r in rows if r["type"] == "income")
    total_expense = sum(r["amount"] for r in rows if r["type"] == "expense")
    balance = total_income - total_expense

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=20
    )

    # Title
    elements.append(Paragraph(f"Money Tracker - Financial Report", title_style))
    elements.append(Paragraph(f"Generated on {datetime.now().strftime('%B %d, %Y at %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 0.2 * inch))

    # Summary table
    summary_data = [
        ["Total Income", "Total Expense", "Net Balance"],
        [f"Rp {total_income:,.0f}", f"Rp {total_expense:,.0f}", f"Rp {balance:,.0f}"]
    ]
    summary_table = Table(summary_data, colWidths=[2 * inch, 2 * inch, 2 * inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ecf0f1')),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3 * inch))

    # Transactions table
    table_data = [["Date", "Type", "Category", "Amount", "Currency", "Note"]]
    for row in rows:
        table_data.append([
            row["date"],
            row["type"].capitalize(),
            row["category"],
            f"Rp {row['amount']:,.0f}",
            row["currency"],
            (row["note"] or "")[:50]
        ])

    transaction_table = Table(table_data, colWidths=[1.0 * inch, 0.8 * inch, 1.2 * inch, 1.2 * inch, 0.8 * inch, 2.2 * inch])
    transaction_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(transaction_table)

    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        download_name=f"transactions_{g.username}.pdf",
        as_attachment=True,
        mimetype="application/pdf"
    )